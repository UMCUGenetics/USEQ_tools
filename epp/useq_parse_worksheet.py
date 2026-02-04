"""Module for parsing and updating LIMS artifacts from worksheet data."""

import sys
import xml.etree.ElementTree as ET
from io import BytesIO
from typing import Dict, List, Tuple, Optional, Any, TextIO

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from genologics.entities import Artifact, Step, Sample, Container
from genologics.lims import Lims

from config import Config


# Column name constants
COLUMN_NAMES = {
    'nr': None,
    'container name': None,
    'sample': None,
    'pre conc ng/ul': None,
    'RIN': None,
    'barcode nr': None,
    'post conc ng/ul': None,
    'size': None
}

# Step name constants
STEP_ISOLATION = 'USEQ - Isolation'
STEP_ISOLATION_V2 = 'USEQ - Isolation v2'
STEP_PRE_LIBPREP_QC = 'USEQ - Pre LibPrep QC'
STEP_LIBPREP_ILLUMINA = 'USEQ - LibPrep Illumina'
STEP_LIBPREP_NANOPORE = 'USEQ - LibPrep Nanopore'
STEP_POST_LIBPREP_QC = 'USEQ - Post LibPrep QC'

# UDF name constants
UDF_CONCENTRATION = 'Concentration Qubit QC (DNA) 5.0 (ng/ul)'
UDF_RIN = 'RIN'
UDF_AVERAGE_LENGTH = 'Average length (bp)'
UDF_PLATFORM = 'Platform'

# Platform constants
PLATFORM_NIMAGEN = '60 SNP NimaGen panel'


def _get_barcode_set(first_sample: Sample, mode: str) -> Optional[Dict[str, str]]:
    """
    Determine which barcode set to use based on platform and mode.

    Args:
        first_sample (Sample): The first sample from the step
        mode (str): Processing mode ('illumina' or 'ont')

    Returns:
        Dictionary of barcodes or None
    """
    if mode == 'illumina':
        if first_sample.udf[UDF_PLATFORM] == PLATFORM_NIMAGEN:
            return Config.NIMAGEN_BARCODES
        return Config.UMI_BARCODES
    elif mode == 'ont':
        return Config.ONT_BARCODES
    return None


def _get_worksheet_columns(sample_worksheet: Worksheet) -> Dict[str, Optional[int]]:
    """
    Extract column indices from worksheet header.

    Args:
        sample_worksheet (Worksheet): The worksheet to extract columns from

    Returns:
        Dictionary mapping column names to their indices

    Raises:
        SystemExit: If 'sample' column is not found
    """
    columns = COLUMN_NAMES.copy()

    for col in sample_worksheet.iter_cols(max_row=1, max_col=sample_worksheet.max_column):
        if col[0].value in columns:
            columns[col[0].value] = col[0].column - 1

    if not columns['sample']:
        sys.exit('ERROR: No "sample" column found.')

    return columns


def _validate_column_exists(columns: Dict[str, Optional[int]], column_name: str):
    """
    Validate that a required column exists.

    Args:
        columns (Dict[str, Optional[int]]): Dictionary of column indices
        column_name (str): Name of the column to validate

    Raises:
        SystemExit: If column is not found
    """
    if not columns[column_name]:
        sys.exit(f'ERROR: No "{column_name}" column found.')


def _get_cell_value(row_cells: Tuple[Cell, ...], columns: Dict[str, Optional[int]], column_name: str, row_nr: int, required: bool = True) -> Any:
    """
    Get value from a cell, with optional validation.

    Args:
        row_cells (Tuple[Cell, ...]): Row cells to extract from
        columns (Dict[str, Optional[int]]): Dictionary of column indices
        column_name (str): Name of the column
        row_nr (int): Current row number for error reporting
        required (bool): Whether the value is required

    Returns:
        Cell value if found

    Raises:
        SystemExit: If required value is missing
    """
    value = row_cells[columns[column_name]].value

    if required and not value:
        sys.exit(f'ERROR: No "{column_name}" found at row {row_nr}.')

    return value


def _parse_isolation_data(row_cells: Tuple[Cell, ...], columns: Dict[str, Optional[int]], row_nr: int, current_step: str) -> Dict[str, Any]:
    """
    Parse data specific to isolation steps.

    Args:
        row_cells (Tuple[Cell, ...]): Row cells to parse
        columns (Dict[str, Optional[int]]): Dictionary of column indices
        row_nr (int): Current row number
        current_step (str): Name of current step

    Returns:
        Dictionary of parsed sample data
    """
    sample = {}

    if current_step in (STEP_ISOLATION, STEP_ISOLATION_V2):
        _validate_column_exists(columns, 'pre conc ng/ul')
        sample['pre conc ng/ul'] = _get_cell_value(
            row_cells, columns, 'pre conc ng/ul', row_nr
        )

        container_name = row_cells[columns['container name']].value
        if container_name:
            sample['container name'] = container_name

    return sample


def _parse_pre_libprep_qc_data(row_cells: Tuple[Cell, ...], columns: Dict[str, Optional[int]], row_nr: int) -> Dict[str, Any]:
    """
    Parse data specific to pre-LibPrep QC step.

    Args:
        row_cells (Tuple[Cell, ...]): Row cells to parse
        columns (Dict[str, Optional[int]]): Dictionary of column indices
        row_nr (int): Current row number

    Returns:
        Dictionary of parsed sample data
    """
    sample = {}
    _validate_column_exists(columns, 'pre conc ng/ul')
    sample['pre conc ng/ul'] = _get_cell_value(
        row_cells, columns, 'pre conc ng/ul', row_nr
    )

    if columns['RIN']:
        rin_value = row_cells[columns['RIN']].value
        if rin_value:
            sample['RIN'] = rin_value

    return sample


def _parse_libprep_data(row_cells: Tuple[Cell, ...], columns: Dict[str, Optional[int]], row_nr: int, barcode_set: Optional[Dict[str, str]]) -> Dict[str, Any]:
    """
    Parse data specific to LibPrep steps.

    Args:
        row_cells (Tuple[Cell, ...]): Row cells to parse
        columns (Dict[str, Optional[int]]): Dictionary of column indices
        row_nr (int): Current row number
        barcode_set (Optional[Dict[str, str]]): Dictionary of available barcodes

    Returns:
        Dictionary of parsed sample data

    Raises:
        SystemExit: If barcode is invalid
    """
    sample = {}
    _validate_column_exists(columns, 'barcode nr')

    barcode_nr = row_cells[columns['barcode nr']].value
    if barcode_nr and barcode_set and barcode_nr in barcode_set:
        sample['barcode'] = barcode_set[barcode_nr]
    else:
        sys.exit(f'ERROR: No valid "barcode nr" found at row {row_nr}.')

    return sample


def _parse_post_libprep_qc_data(row_cells: Tuple[Cell, ...], columns: Dict[str, Optional[int]], row_nr: int) -> Dict[str, Any]:
    """
    Parse data specific to post-LibPrep QC step.

    Args:
        row_cells (Tuple[Cell, ...]): Row cells to parse
        columns (Dict[str, Optional[int]]): Dictionary of column indices
        row_nr (int): Current row number

    Returns:
        Dictionary of parsed sample data
    """
    sample = {}
    _validate_column_exists(columns, 'post conc ng/ul')
    sample['post conc ng/ul'] = _get_cell_value(
        row_cells, columns, 'post conc ng/ul', row_nr
    )

    _validate_column_exists(columns, 'size')
    sample['size'] = _get_cell_value(row_cells, columns, 'size', row_nr)

    return sample


def _parse_samples_from_worksheet(sample_worksheet: Worksheet, columns: Dict[str, Optional[int]], current_step: str, barcode_set: Optional[Dict[str, str]]) -> Dict[str, Dict[str, Any]]:
    """
    Parse all samples from the worksheet.

    Args:
        sample_worksheet (Worksheet): Worksheet containing sample data
        columns (Dict[str, Optional[int]]): Dictionary of column indices
        current_step (str): Name of current processing step
        barcode_set (Optional[Dict[str, str]]): Dictionary of available barcodes

    Returns:
        Dictionary mapping sample names to their data
    """
    samples = {}
    row_nr = 1

    for row_cells in sample_worksheet.iter_rows(min_row=2, max_row=sample_worksheet.max_row):
        if not row_cells[columns['sample']].value:
            continue

        sample = {}

        # Parse data based on current step
        if current_step in (STEP_ISOLATION, STEP_ISOLATION_V2):
            sample.update(_parse_isolation_data(row_cells, columns, row_nr, current_step))

        if current_step == STEP_PRE_LIBPREP_QC:
            sample.update(_parse_pre_libprep_qc_data(row_cells, columns, row_nr))

        if current_step in (STEP_LIBPREP_ILLUMINA, STEP_LIBPREP_NANOPORE):
            sample.update(_parse_libprep_data(row_cells, columns, row_nr, barcode_set))

        if current_step == STEP_POST_LIBPREP_QC:
            sample.update(_parse_post_libprep_qc_data(row_cells, columns, row_nr))

        sample_name = str(row_cells[columns['sample']].value)
        if sample_name not in samples:
            samples[sample_name] = sample

        row_nr += 1

    return samples


def _update_artifact_concentration(artifact: Artifact, sample_info: Dict[str, Any], concentration_key: str):
    """
    Update artifact concentration UDF if not already set.

    Args:
        artifact (Artifact): LIMS Artifact to update
        sample_info (Dict[str, Any]): Dictionary containing sample information
        concentration_key (str): Key for concentration value in sample_info
    """
    if concentration_key in sample_info and UDF_CONCENTRATION not in artifact.udf:
        artifact.udf[UDF_CONCENTRATION] = float(sample_info[concentration_key])


def _update_artifact_container(artifact: Artifact, sample_info: Dict[str, Any], first_sample: Sample, containers_to_update: List[Container]):
    """
    Update artifact container name if specified.

    Args:
        artifact (Artifact): LIMS Artifact to update
        sample_info (Dict[str, Any]): Dictionary containing sample information
        first_sample (Sample): LIMS Sample
        containers_to_update (List[Container]): List to append updated containers to
    """
    if 'container name' in sample_info:
        container: Container = artifact.location[0]
        container.name = sample_info['container name']
        containers_to_update.append(container)

        # Update artifact name for NimaGen platform
        if (first_sample.udf[UDF_PLATFORM] == PLATFORM_NIMAGEN and
                sample_info['container name'] not in artifact.name):
            artifact.name = f"{sample_info['container name']}-{artifact.samples[0].name}"


def _update_artifact_rin(artifact: Artifact, sample_info: Dict[str, Any]):
    """
    Update artifact RIN UDF if not already set.

    Args:
        artifact (Artifact): Artifact to update
        sample_info (Dict[str, Any]): Dictionary containing sample information
    """
    if 'RIN' in sample_info and UDF_RIN not in artifact.udf:
        artifact.udf[UDF_RIN] = float(sample_info['RIN'])


def _update_artifact_barcode(artifact: Artifact, sample_info: Dict[str, Any], artifact_sample: Sample, project_id: str, output_file: TextIO):
    """
    Add barcode to artifact and log the operation.

    Args:
        artifact (Artifact): LIMS Artifact to update
        sample_info (Dict[str, Any]): Dictionary containing sample information
        artifact_sample (Sample): LIMS Sample associated with the artifact
        project_id (str): Project ID for logging
        output_file (TextIO): File to write log output
    """
    if 'barcode' in sample_info:
        reagent_label = ET.SubElement(artifact.root, 'reagent-label')
        reagent_label.set('name', sample_info['barcode'])
        output_file.write(
            f"Barcode added to: {project_id}\t{artifact_sample.name}\t"
            f"{sample_info['barcode']}\n"
        )


def _update_artifact_size(artifact: Artifact, sample_info: Dict[str, Any]):
    """
    Update artifact size UDF if not already set.

    Args:
        artifact (Artifact): LIMS Artifact to update
        sample_info (Dict[str, Any]): Dictionary containing sample information
    """
    if 'size' in sample_info and UDF_AVERAGE_LENGTH not in artifact.udf:
        artifact.udf[UDF_AVERAGE_LENGTH] = int(sample_info['size'])


def _update_artifacts(step: Step, samples: Dict[str, Dict[str, Any]], first_sample: Sample, output_file: TextIO) -> Tuple[List[Artifact], List[Container]]:
    """
    Update all artifacts based on parsed sample data.

    Args:
        step (Step): LIMS step object
        samples (Dict[str, Dict[str, Any]]): Dictionary of parsed sample data
        first_sample (Sample): LIMS sample
        output_file (TextIO): File to write log output

    Returns:
        Tuple of (artifacts_to_update, containers_to_update)
    """
    artifacts_to_update = []
    containers_to_update = []

    for io_map in step.details.input_output_maps:
        if io_map[1]['output-generation-type'] != 'PerInput':
            continue

        artifact = io_map[1]['uri']
        artifact_sample = artifact.samples[0]
        sample_info = samples[str(artifact_sample.name)]
        project_id = artifact_sample.project.id

        # Update various artifact properties
        _update_artifact_concentration(artifact, sample_info, 'pre conc ng/ul')
        _update_artifact_container(artifact, sample_info, first_sample, containers_to_update)
        _update_artifact_rin(artifact, sample_info)
        _update_artifact_barcode(artifact, sample_info, artifact_sample, project_id, output_file)
        _update_artifact_concentration(artifact, sample_info, 'post conc ng/ul')
        _update_artifact_size(artifact, sample_info)

        artifacts_to_update.append(artifact)

    return artifacts_to_update, containers_to_update


def parse(lims: Lims, step_uri: str, aid: str, output_file: TextIO, mode: str):
    """
    Parse worksheet data and update LIMS artifacts.

    Args:
        lims (Lims): LIMS connection object
        step_uri (str): URI of the processing step
        aid (str): Artifact ID of the worksheet
        output_file (TextIO): File object to write output logs
        mode (str): Processing mode ('illumina' or 'ont')
    """
    # Get worksheet artifact and content
    worksheet_artifact = Artifact(lims, id=aid)
    worksheet_id = worksheet_artifact.files[0].id

    step = Step(lims, uri=step_uri)
    current_step = step.configuration.name

    first_sample = step.details.input_output_maps[0][0]['uri'].samples[0]
    barcode_set = _get_barcode_set(first_sample, mode)

    # Load worksheet
    content = lims.get_file_contents(id=worksheet_id).read()
    workbook = load_workbook(filename=BytesIO(content))
    sample_worksheet = workbook['Samples']

    # Parse worksheet structure and data
    columns = _get_worksheet_columns(sample_worksheet)
    samples = _parse_samples_from_worksheet(
        sample_worksheet, columns, current_step, barcode_set
    )

    # Update artifacts
    artifacts_to_update, containers_to_update = _update_artifacts(
        step, samples, first_sample, output_file
    )

    # Commit changes to LIMS
    lims.put_batch(artifacts_to_update)
    lims.put_batch(containers_to_update)


def run(lims: Lims, step_uri: str, aid: str, output_file: TextIO, mode: str):
    """Entry point for the script.

    Args:
        lims (Lims): LIMS connection object
        step_uri (str): URI of the processing step
        aid (str): Artifact ID of the worksheet
        output_file (TextIO): File object to write output logs
        mode (str): Processing mode ('illumina' or 'ont')
    """
    parse(lims, step_uri, aid, output_file, mode)
